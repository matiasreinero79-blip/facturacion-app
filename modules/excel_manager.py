from pathlib import Path
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# Canonical column order: (field_key, display_header)
# NOTE: cuit and tipo_factura added after existing columns for backward compatibility.
# Old Excel files (missing these columns) will simply show empty values.
COLUMNS: List[tuple] = [
    ("numero_cuenta",   "Número de Cuenta"),
    ("numero_cliente",  "Número de Cliente"),
    ("empresa",         "Empresa"),
    ("numero_factura",  "Número de Factura"),
    ("cuit",            "CUIT"),
    ("tipo_factura",    "Tipo de Factura"),
    ("importe",         "Importe"),
    ("fecha_emision",   "Fecha de Emisión"),
    ("fecha_envio",     "Fecha de Envío"),
    ("estado_pago",     "Estado de Pago"),
    ("archivo",         "Archivo PDF"),
    ("ruta",            "Ruta del Archivo"),
    ("fecha_carga",     "Fecha de Carga"),
]

_COL_WIDTHS = [18, 18, 28, 22, 20, 14, 14, 16, 16, 14, 32, 55, 22]

_FILL_HEADER    = PatternFill("solid", fgColor="1F4E79")
_FILL_EVEN      = PatternFill("solid", fgColor="DCE6F1")
_FILL_ODD       = PatternFill("solid", fgColor="FFFFFF")
_FILL_PAGA      = PatternFill("solid", fgColor="C6EFCE")
_FILL_PENDIENTE = PatternFill("solid", fgColor="FFEB9C")
_FILL_VENCIDO   = PatternFill("solid", fgColor="FFC7CE")  # rojo suave

_FONT_HEADER    = Font(color="FFFFFF", bold=True, size=11, name="Arial")
_FONT_PAGA      = Font(color="276221", name="Arial")
_FONT_PENDIENTE = Font(color="9C5700", name="Arial")
_FONT_VENCIDO   = Font(color="9C0006", name="Arial")

_THIN  = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ESTADO_COL_IDX = next(i for i, (k, _) in enumerate(COLUMNS) if k == "estado_pago")


class ExcelManager:

    def __init__(self, excel_path: str = "control_facturas.xlsx"):
        self.excel_path = Path(excel_path)

    # ─── public API ─────────────────────────────────────────────────────────

    def add_invoice(self, data: Dict) -> bool:
        """
        Append *data* as a new row.
        Returns False (and does nothing) when a duplicate is detected.
        """
        empresa = (data.get("empresa") or "").strip()
        numero  = (data.get("numero_factura") or "").strip()

        if empresa and numero and self.invoice_exists(empresa, numero):
            return False

        if not data.get("fecha_carga"):
            data["fecha_carga"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        wb = self._load_or_create()
        ws = wb.active
        row_idx = self._next_row(ws)
        self._write_row(ws, row_idx, data)
        wb.save(str(self.excel_path))
        return True

    def invoice_exists(self, empresa: str, numero_factura: str) -> bool:
        """Check by empresa + numero_factura (case-insensitive)."""
        emp_lo = empresa.strip().lower()
        num_lo = numero_factura.strip().lower()
        for inv in self.get_all_invoices():
            if (inv.get("empresa", "").lower() == emp_lo and
                    inv.get("numero_factura", "").lower() == num_lo):
                return True
        return False

    def get_all_invoices(self) -> List[Dict]:
        if not self.excel_path.exists():
            return []
        wb = openpyxl.load_workbook(str(self.excel_path), read_only=True, data_only=True)
        ws = wb.active
        invoices = []

        # Read actual headers from row 1 to handle old files gracefully
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_map = {str(h).strip(): i for i, h in enumerate(header_row) if h}

        # Build a mapping: column_key → column_index_in_file
        col_indices = {}
        for key, header in COLUMNS:
            if header in header_map:
                col_indices[key] = header_map[header]
            # If header not found (old file), key will map to None → empty string

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None for c in row):
                continue
            inv = {}
            for key, _ in COLUMNS:
                idx = col_indices.get(key)
                val = row[idx] if (idx is not None and idx < len(row)) else None
                inv[key] = "" if val is None else str(val)
            invoices.append(inv)
        wb.close()
        return invoices

    def get_pending_invoices(self) -> List[Dict]:
        """Devuelve facturas con estado Pendiente o Vencido (ambas requieren atención)."""
        return [
            i for i in self.get_all_invoices()
            if i.get("estado_pago", "").strip().lower() in ("pendiente", "vencido")
        ]

    def update_invoice_status(self, empresa: str, numero_factura: str, new_status: str) -> bool:
        if not self.excel_path.exists():
            return False
        wb = openpyxl.load_workbook(str(self.excel_path))
        ws = wb.active

        # Read headers to find column positions dynamically
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        h_map = {str(h).strip(): i + 1 for i, h in enumerate(header_row) if h}

        emp_col = h_map.get("Empresa")
        num_col = h_map.get("Número de Factura")
        est_col = h_map.get("Estado de Pago")

        if not (emp_col and num_col and est_col):
            wb.close()
            return False

        found = False
        for row in ws.iter_rows(min_row=2):
            emp_val = str(row[emp_col - 1].value or "").strip().lower()
            num_val = str(row[num_col - 1].value or "").strip().lower()
            if emp_val == empresa.strip().lower() and num_val == numero_factura.strip().lower():
                cell = row[est_col - 1]
                cell.value = new_status
                if new_status.strip().lower() == "paga":
                    cell.fill = _FILL_PAGA
                    cell.font = _FONT_PAGA
                elif new_status.strip().lower() == "vencido":
                    cell.fill = _FILL_VENCIDO
                    cell.font = _FONT_VENCIDO
                else:  # Pendiente (default)
                    cell.fill = _FILL_PENDIENTE
                    cell.font = _FONT_PENDIENTE
                found = True
                break

        if found:
            wb.save(str(self.excel_path))
        return found

    # ─── internals ──────────────────────────────────────────────────────────

    def _load_or_create(self) -> openpyxl.Workbook:
        if self.excel_path.exists():
            return self._migrate_if_needed(openpyxl.load_workbook(str(self.excel_path)))
        return self._new_workbook()

    def _migrate_if_needed(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """Add missing columns (CUIT, Tipo de Factura) to existing workbooks."""
        ws = wb.active
        existing_headers = [
            str(ws.cell(1, c).value or "").strip()
            for c in range(1, ws.max_column + 1)
        ]
        expected_headers = [h for _, h in COLUMNS]

        if existing_headers == expected_headers:
            return wb  # already up to date

        # Rebuild header row with any missing columns inserted at correct positions
        # For simplicity: append missing columns at the end of existing ones
        for col_key, col_header in COLUMNS:
            if col_header not in existing_headers:
                new_col = ws.max_column + 1
                cell = ws.cell(1, new_col, value=col_header)
                cell.fill      = _FILL_HEADER
                cell.font      = _FONT_HEADER
                cell.border    = _BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.column_dimensions[get_column_letter(new_col)].width = 18

        return wb

    def _new_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"
        ws.freeze_panes = "A2"

        for col_idx, (_, header) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = _FILL_HEADER
            cell.font      = _FONT_HEADER
            cell.border    = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28
        for i, w in enumerate(_COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        return wb

    def _next_row(self, ws) -> int:
        row = ws.max_row + 1
        # If the sheet only has the header, start at row 2
        if row == 2 and ws.cell(2, 1).value is None:
            row = 2
        return row

    def _write_row(self, ws, row_idx: int, data: Dict) -> None:
        row_fill = _FILL_EVEN if row_idx % 2 == 0 else _FILL_ODD

        # Build column map from current sheet headers
        header_row = [str(ws.cell(1, c).value or "").strip()
                      for c in range(1, ws.max_column + 1)]
        h_map = {h: i + 1 for i, h in enumerate(header_row) if h}

        for key, header in COLUMNS:
            col_idx = h_map.get(header)
            if col_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col_idx, value=data.get(key, "") or "")
            cell.border    = _BORDER
            cell.alignment = Alignment(vertical="center")
            cell.fill      = row_fill

        # Colour the estado_pago cell
        est_col = h_map.get("Estado de Pago")
        if est_col:
            estado_cell = ws.cell(row=row_idx, column=est_col)
            if (data.get("estado_pago") or "").lower() == "paga":
                estado_cell.fill = _FILL_PAGA
                estado_cell.font = _FONT_PAGA
            else:
                estado_cell.fill = _FILL_PENDIENTE
                estado_cell.font = _FONT_PENDIENTE
